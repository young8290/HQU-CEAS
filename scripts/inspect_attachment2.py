import struct
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
XLS_PATH = ROOT / "附件2.2024-2025学年院级奖学金、优秀学生干部、优秀学生汇总表【含填报说明】.xls"
XLSX_TEMPLATE = ROOT / "templates" / "附件2_院奖荣誉汇总_template.xlsx"

END_OF_CHAIN = 0xFFFFFFFE
FREE_SECTOR = 0xFFFFFFFF


def u16(data, offset):
    return struct.unpack_from("<H", data, offset)[0]


def u32(data, offset):
    return struct.unpack_from("<I", data, offset)[0]


def i32(data, offset):
    return struct.unpack_from("<i", data, offset)[0]


def sector_offset(sid, sector_size):
    return (sid + 1) * sector_size


def parse_cfb(path):
    data = path.read_bytes()
    sector_size = 1 << u16(data, 30)
    first_dir = u32(data, 48)
    mini_cutoff = u32(data, 56)
    difat = [u32(data, 76 + index * 4) for index in range(109)]
    fat_sector_ids = [sid for sid in difat if sid not in (FREE_SECTOR, END_OF_CHAIN)]

    fat = []
    for sid in fat_sector_ids:
        sector = data[sector_offset(sid, sector_size):sector_offset(sid, sector_size) + sector_size]
        fat.extend(struct.unpack("<" + "I" * (sector_size // 4), sector))

    def read_chain(start):
        chunks = []
        sid = start
        seen = set()
        while sid not in (END_OF_CHAIN, FREE_SECTOR) and sid < len(fat) and sid not in seen:
            seen.add(sid)
            chunks.append(data[sector_offset(sid, sector_size):sector_offset(sid, sector_size) + sector_size])
            sid = fat[sid]
        return b"".join(chunks)

    directory = read_chain(first_dir)
    entries = {}
    root_stream = b""
    for offset in range(0, len(directory), 128):
        entry = directory[offset:offset + 128]
        name_len = u16(entry, 64)
        if name_len < 2:
            continue
        name = entry[:name_len - 2].decode("utf-16le", errors="ignore")
        entry_type = entry[66]
        start = u32(entry, 116)
        size = struct.unpack_from("<Q", entry, 120)[0]
        if entry_type == 5:
            root_stream = read_chain(start)[:size]
        elif entry_type == 2 and size >= mini_cutoff:
            entries[name] = read_chain(start)[:size]

    return entries, root_stream


def read_biff_string(data, offset):
    length = u16(data, offset)
    flags = data[offset + 2]
    offset += 3
    rich_runs = u16(data, offset) if flags & 0x08 else 0
    if flags & 0x08:
        offset += 2
    ext_size = u32(data, offset) if flags & 0x04 else 0
    if flags & 0x04:
        offset += 4
    if flags & 0x01:
        raw = data[offset:offset + length * 2]
        text = raw.decode("utf-16le", errors="ignore")
        offset += length * 2
    else:
        raw = data[offset:offset + length]
        text = raw.decode("latin1", errors="ignore")
        offset += length
    offset += rich_runs * 4 + ext_size
    return text, offset


def decode_rk(raw):
    mult100 = raw & 0x01
    is_int = raw & 0x02
    if is_int:
        value = i32(struct.pack("<I", raw & 0xFFFFFFFC), 0) >> 2
    else:
        value = struct.unpack("<d", struct.pack("<I", raw & 0xFFFFFFFC) + b"\x00\x00\x00\x00")[0]
    return value / 100 if mult100 else value


def parse_workbook_stream(stream):
    records = []
    position = 0
    while position + 4 <= len(stream):
        opcode = u16(stream, position)
        length = u16(stream, position + 2)
        payload = stream[position + 4:position + 4 + length]
        records.append((position, opcode, payload))
        position += 4 + length

    sheets = []
    for _, opcode, payload in records:
        if opcode != 0x0085 or len(payload) < 8:
            continue
        bof_offset = u32(payload, 0)
        name_len = payload[6]
        flags = payload[7]
        raw = payload[8:]
        if flags & 0x01:
            name = raw[:name_len * 2].decode("utf-16le", errors="ignore")
        else:
            name = raw[:name_len].decode("latin1", errors="ignore")
        sheets.append({"name": name, "offset": bof_offset, "cells": {}, "merged": []})
    sheets.sort(key=lambda item: item["offset"])

    sst = []
    for index, (_, opcode, payload) in enumerate(records):
        if opcode != 0x00FC or len(payload) < 8:
            continue
        combined = bytearray(payload)
        next_index = index + 1
        while next_index < len(records) and records[next_index][1] == 0x003C:
            combined.extend(records[next_index][2])
            next_index += 1
        data = bytes(combined)
        total = u32(data, 4)
        offset = 8
        for _ in range(total):
            if offset + 3 > len(data):
                break
            text, offset = read_biff_string(data, offset)
            sst.append(text)
        break

    def sheet_for_position(pos):
        result = None
        for sheet in sheets:
            if sheet["offset"] <= pos:
                result = sheet
            else:
                break
        return result

    for position, opcode, payload in records:
        sheet = sheet_for_position(position)
        if not sheet:
            continue
        if opcode == 0x00FD and len(payload) >= 10:
            row, col = u16(payload, 0), u16(payload, 2)
            sst_index = u32(payload, 6)
            sheet["cells"][(row + 1, col + 1)] = sst[sst_index] if sst_index < len(sst) else ""
        elif opcode == 0x0203 and len(payload) >= 14:
            row, col = u16(payload, 0), u16(payload, 2)
            sheet["cells"][(row + 1, col + 1)] = struct.unpack_from("<d", payload, 6)[0]
        elif opcode == 0x027E and len(payload) >= 10:
            row, col = u16(payload, 0), u16(payload, 2)
            sheet["cells"][(row + 1, col + 1)] = decode_rk(u32(payload, 6))
        elif opcode == 0x00BD and len(payload) >= 6:
            row = u16(payload, 0)
            first_col = u16(payload, 2)
            last_col = u16(payload, len(payload) - 2)
            offset = 4
            for col in range(first_col, last_col + 1):
                if offset + 6 > len(payload) - 2:
                    break
                sheet["cells"][(row + 1, col + 1)] = decode_rk(u32(payload, offset + 2))
                offset += 6
        elif opcode == 0x0204 and len(payload) >= 8:
            row, col = u16(payload, 0), u16(payload, 2)
            text, _ = read_biff_string(payload, 6)
            sheet["cells"][(row + 1, col + 1)] = text
        elif opcode == 0x00E5 and len(payload) >= 2:
            count = u16(payload, 0)
            for index in range(count):
                offset = 2 + index * 8
                if offset + 8 > len(payload):
                    break
                first_row, last_row, first_col, last_col = struct.unpack_from("<HHHH", payload, offset)
                sheet["merged"].append((first_row + 1, first_col + 1, last_row + 1, last_col + 1))

    return sheets


def print_sheet(sheet, max_rows=18, max_cols=24):
    print(f"工作表：{sheet['name']}")
    if sheet["merged"]:
        merged = [
            f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
            for r1, c1, r2, c2 in sheet["merged"][:20]
        ]
        print("合并单元格：" + "，".join(merged))
    for row in range(1, max_rows + 1):
        values = []
        for col in range(1, max_cols + 1):
            value = sheet["cells"].get((row, col), "")
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            values.append(str(value).replace("\n", " ").strip())
        if any(values):
            print(f"{row:02d} | " + " | ".join(values))
    print()


def inspect_xls():
    streams, _ = parse_cfb(XLS_PATH)
    workbook = streams.get("Workbook") or streams.get("Book")
    if not workbook:
        raise RuntimeError("Workbook stream not found")
    print(f"原始 XLS：{XLS_PATH.name}")
    for sheet in parse_workbook_stream(workbook):
        print_sheet(sheet)


def inspect_xlsx():
    workbook = load_workbook(XLSX_TEMPLATE, data_only=False)
    print(f"当前 XLSX 模板：{XLSX_TEMPLATE.name}")
    for worksheet in workbook.worksheets:
        sheet = {
            "name": worksheet.title,
            "cells": {},
            "merged": [],
        }
        for row in worksheet.iter_rows(min_row=1, max_row=18, min_col=1, max_col=24):
            for cell in row:
                if cell.value is not None:
                    sheet["cells"][(cell.row, cell.column)] = cell.value
        for merged in worksheet.merged_cells.ranges:
            sheet["merged"].append((merged.min_row, merged.min_col, merged.max_row, merged.max_col))
        print_sheet(sheet)


if __name__ == "__main__":
    inspect_xls()
    inspect_xlsx()
